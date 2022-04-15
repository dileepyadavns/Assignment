
#imported libraries
import openpyxl
import threading

path="multi_threading_activity.xlsx" #path of xlsx file
obj = openpyxl.load_workbook(path)
sheet1 = obj.active

#Read each column  and row of xl file and added to list
mat=[]
for i in range(1, sheet1.max_row+1):
  row=[]
  for j in range(1, sheet1.max_column+1):
    cell_obj = sheet1.cell(row=i, column=j)
    row.append(cell_obj.value)
  mat.append(row)
print(mat)  


mat.pop() #Got null list at so popped
mat.pop()#Got null list at so popped

#Replaced all the strings with value 0
new=mat[1:]
for i in range(len(new)):
  for j in range(2):
    if  type(new[i][j])==str:
      print(new[i][j])
      new[i][j]=0
    if  type(new[i][j])==None:
      print(new[i][j])
      new[i][j]=0

print(new)      
print(len(new))

#calculated the sum of columns by adding corresonding rows
result=[]
def sum_of_two_rows(new):
  sumMat=[]
  for i in range(len(new)):
    sumMat.append(new[i][0]+new[i][1])
  result.append(sumMat) 
  print(sumMat)


#Found the difference of columns by subbtracting corresponding rows
def diff_of_tow_rows(new):
  diffMat=[]
  for i in range(len(new)):
    diffMat.append(new[i][0]-new[i][1])
  result.append(diffMat)
  print(diffMat)

t1=threading.Thread(target=sum_of_two_rows,args=(new,))
t2=threading.Thread(target=diff_of_tow_rows,args=(new,))
t1.start()
t2.start()
a=t1.join()
b=t2.join()
print(result) #it contains list of sum and diffference of rows
col1=result[0]
print(col1) #it gives sum

col2=result[1]
print(col2) # it gives difference

